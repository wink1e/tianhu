import time

import openpyxl

import formatted_print
import requests
import json
import datetime
from tqdm import tqdm
import toml

config = toml.load('config.toml')


def fetch_data_from_tziakcha(record_id, session):
    url = r"https://tziakcha.xyz/_qry/game/"
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        'Content-Type': 'text/plain;charset=UTF-8',
        'Cookie': '__p=ee469c7bc00d96ca79a40ccf62adf02506bc473d7a30a535e190aeb3ab60f98a',
        'Origin': 'https://tziakcha.xyz',
        'Referer': 'https://tziakcha.xyz/game/?id=10045',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Google Chrome";v="125", "Chromium";v="125", "Not.A/Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Linux"'
    }
    data = f"?id={record_id}"
    url += data
    attempt = 1
    while True:
        try:
            response = session.post(url, headers=headers, timeout=5)
            if response.status_code == 404:
                return record_id, 404
            response.raise_for_status()
            return record_id, response.json()
        except requests.exceptions.RequestException as err:
            print(url, data, err)
            time.sleep(4 * attempt)
            print(f'retry for {attempt} time')
            attempt += 1
            if attempt == 5:
                return  record_id, 404


class OutOfDataException(Exception):
    def __init__(self, message, error_code):
        super().__init__(message)
        self.error_code = error_code


def convert_data(record_json):
    global config
    if "天和" not in record_json['title']:
        return None
    if "elo0" not in record_json.keys():
        raise OutOfDataException("Out of the paipu data", 500)
    start_time = datetime.datetime.fromtimestamp(record_json["start_time"]/1000).strftime("%Y-%m-%d %H:%M:%S")
    end_time = datetime.datetime.fromtimestamp(record_json["finish_time"]/1000).strftime("%Y-%m-%d %H:%M:%S")
    record = {}
    result = {}
    uid = []
    for i in range(4):
        result[record_json[f"name{i}"]] = record_json[f"score{i}"]
        uid.append(record_json[f"uid{i}"])
    record["result"] = result
    record["title"] = record_json["title"]
    record["start_time"] = start_time
    record["end_time"] = end_time
    record["game_id"] = record_json["id"]
    record["uid"] = uid
    return record


def write_excel(out, result_id):
    file_path = config["excel_file"]["excel_file_path"]
    file = openpyxl.load_workbook(file_path)
    sheet = file["games"]
    row = result_id + 2
    sheet.cell(row, 1, result_id)
    sheet.cell(row, 2, out["game_id"])
    sheet.cell(row, 3, out["title"])
    sheet.cell(row, 4, out["start_time"])
    sheet.cell(row, 5, out["end_time"])
    index = 6
    i = 0
    for keys, values in out["result"].items():
        sheet.cell(row, index, keys)
        sheet.cell(row, index+1, values)
        sheet.cell(row, index+2, out["uid"][i])
        i += 1
        index += 3
    file.save(file_path)


def update():
    global config
    start_id = config["id"]["record_id"]
    result_id = config["excel_file"]["result_id"]
    session = requests.Session()
    while True:
        try:
            _, json_temp = fetch_data_from_tziakcha(start_id, session)
            if json_temp == 404:
                raise OutOfDataException
            out = convert_data(json_temp)
            if out is None:
                start_id += 1
                continue
            formatted_print.custom_pretty_print(out)
            write_excel(out, result_id=result_id)
            start_id += 1
            result_id += 1
        except OutOfDataException as err:
            print("All of the data are updated")
            config["id"]["record_id"] = start_id
            config["excel_file"]["result_id"] = result_id
            now = datetime.datetime.now()
            config["date"]["ld1"] = now.date()
            config["date"]["lt1"] = now.time().strftime("%H:%M:%S")
            with open('config.toml', "w", encoding='utf-8') as f:
                toml.dump(config, f,)
            break


if __name__ == "__main__":
    update()
    # session = requests.Session()
    # _, json_temp = fetch_data_from_tziakcha(record_id=115636, session=session)
    # formatted_print.custom_pretty_print(json_temp)